# (imports permanecem os mesmos)
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from ttkthemes import ThemedTk
import sqlite3
import os
import csv
from datetime import datetime
from dateutil.relativedelta import relativedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class GestaoEditaisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestão de Editais")
        self.root.geometry("")
        self.root.minsize(850, 600)
        self.root.resizable(True, True)

        self.db_path = os.path.join(os.path.dirname(__file__), 'gestao_editais.db')

        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        self.tab_edital = ttk.Frame(self.notebook)
        self.tab_modalidades = ttk.Frame(self.notebook)
        self.tab_bolsistas = ttk.Frame(self.notebook)
        self.tab_consulta = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_edital, text="Cadastro de Edital")
        self.notebook.add(self.tab_modalidades, text="Modalidades")
        self.notebook.add(self.tab_bolsistas, text="Bolsistas")
        self.notebook.add(self.tab_consulta, text="Consulta")

        self.criar_formulario_edital()
        self.criar_formulario_modalidades()
        self.criar_formulario_bolsistas()
        self.criar_aba_consulta()

    def converter_data_br_para_iso(self, data_br):
        if not data_br:
            return None
        try:
            data = datetime.strptime(data_br.strip(), "%d/%m/%Y")
            return data.strftime("%Y-%m-%d")
        except ValueError:
            return None

    def converter_data_iso_para_br(self, data_iso):
        if not data_iso:
            return ""
        try:
            data = datetime.strptime(data_iso, "%Y-%m-%d")
            return data.strftime("%d/%m/%Y")
        except ValueError:
            return ""

    def calcular_data_fim_bolsa(self, data_inicio_bolsa_iso, meses):
        try:
            data = datetime.strptime(data_inicio_bolsa_iso, "%Y-%m-%d")
            data_alvo = data + relativedelta(months=meses - 1)
            prox_mes = data_alvo + relativedelta(months=1)
            ultimo_dia = prox_mes - relativedelta(days=1)
            return ultimo_dia.strftime("%Y-%m-%d")
        except Exception as e:
            return None

    # === FORMULÁRIOS (mesmo que antes) ===
    def criar_formulario_edital(self):
        frame = ttk.Frame(self.tab_edital)
        frame.pack(padx=20, pady=20)

        ttk.Label(frame, text="Número do Edital:").grid(row=0, column=0, sticky='w', pady=5)
        self.entry_numero = ttk.Entry(frame, width=40)
        self.entry_numero.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Descrição:").grid(row=1, column=0, sticky='w', pady=5)
        self.entry_descricao = ttk.Entry(frame, width=40)
        self.entry_descricao.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Agência de Fomento:").grid(row=2, column=0, sticky='w', pady=5)
        self.entry_agencia = ttk.Entry(frame, width=40)
        self.entry_agencia.grid(row=2, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Código do Projeto (opcional):").grid(row=3, column=0, sticky='w', pady=5)
        self.entry_projeto_codigo = ttk.Entry(frame, width=40)
        self.entry_projeto_codigo.grid(row=3, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Descrição do Projeto (opcional):").grid(row=4, column=0, sticky='w', pady=5)
        self.entry_projeto_desc = ttk.Entry(frame, width=40)
        self.entry_projeto_desc.grid(row=4, column=1, padx=10, pady=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="Cadastrar", command=self.cadastrar_edital).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Limpar", command=self.limpar_campos_edital).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Sair", command=self.root.destroy).pack(side='left', padx=5)

    def cadastrar_edital(self):
        numero = self.entry_numero.get().strip()
        descricao = self.entry_descricao.get().strip()
        agencia = self.entry_agencia.get().strip()
        cod_projeto = self.entry_projeto_codigo.get().strip() or None
        desc_projeto = self.entry_projeto_desc.get().strip() or None

        if not numero:
            messagebox.showerror("Erro", "O campo 'Número do Edital' é obrigatório.")
            return
        if not agencia:
            messagebox.showerror("Erro", "O campo 'Agência de Fomento' é obrigatório.")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO editais (numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto)
                VALUES (?, ?, ?, ?, ?)
            ''', (numero, descricao, agencia, cod_projeto, desc_projeto))
            conn.commit()
            conn.close()

            messagebox.showinfo("Sucesso", "Edital cadastrado com sucesso!")
            self.limpar_campos_edital()
            self.carregar_editais_para_combo()

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao salvar edital:\n{str(e)}")

    def limpar_campos_edital(self):
        self.entry_numero.delete(0, tk.END)
        self.entry_descricao.delete(0, tk.END)
        self.entry_agencia.delete(0, tk.END)
        self.entry_projeto_codigo.delete(0, tk.END)
        self.entry_projeto_desc.delete(0, tk.END)

    def criar_formulario_modalidades(self):
        frame = ttk.Frame(self.tab_modalidades)
        frame.pack(padx=20, pady=20)

        ttk.Label(frame, text="Selecione o Edital:").grid(row=0, column=0, sticky='w', pady=5)
        self.combo_editais_mod = ttk.Combobox(frame, width=37, state="readonly")
        self.combo_editais_mod.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Nível:").grid(row=1, column=0, sticky='w', pady=5)
        self.combo_nivel_mod = ttk.Combobox(frame, width=37, state="readonly")
        self.combo_nivel_mod['values'] = ("graduação", "mestrado", "doutorado", "pós-doutorado")
        self.combo_nivel_mod.grid(row=1, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Número de Vagas:").grid(row=2, column=0, sticky='w', pady=5)
        self.entry_vagas = ttk.Entry(frame, width=40)
        self.entry_vagas.grid(row=2, column=1, padx=10, pady=5)

        ttk.Label(frame, text="Valor Mensal (R$):").grid(row=3, column=0, sticky='w', pady=5)
        self.entry_valor = ttk.Entry(frame, width=40)
        self.entry_valor.grid(row=3, column=1, padx=10, pady=5)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="Cadastrar", command=self.cadastrar_modalidade).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Limpar", command=self.limpar_campos_modalidade).pack(side='left', padx=5)

        self.carregar_editais_para_combo()

    def carregar_editais_para_combo(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, numero_edital FROM editais ORDER BY id")
            editais = cursor.fetchall()
            conn.close()

            if editais:
                self.combo_editais_mod['values'] = [f"{e[1]}" for e in editais]
                self.combo_editais_mod.set("")
                if hasattr(self, 'combo_editais_bols'):
                    self.combo_editais_bols['values'] = [f"{e[1]}" for e in editais]
                    self.combo_editais_bols.set("")
                if hasattr(self, 'combo_consulta_editais'):
                    self.carregar_editais_consulta()
            else:
                self.combo_editais_mod['values'] = ["Nenhum edital cadastrado"]
                self.combo_editais_mod.set("Nenhum edital cadastrado")
                if hasattr(self, 'combo_editais_bols'):
                    self.combo_editais_bols['values'] = ["Nenhum edital cadastrado"]
                    self.combo_editais_bols.set("Nenhum edital cadastrado")
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar editais:\n{str(e)}")

    def cadastrar_modalidade(self):
        edital_selecionado = self.combo_editais_mod.get()
        nivel = self.combo_nivel_mod.get()
        vagas_str = self.entry_vagas.get().strip()
        valor_str = self.entry_valor.get().strip()

        if edital_selecionado == "Nenhum edital cadastrado" or not edital_selecionado:
            messagebox.showerror("Erro", "Cadastre um edital primeiro.")
            return
        if not nivel:
            messagebox.showerror("Erro", "Selecione o nível da bolsa.")
            return
        if not vagas_str.isdigit():
            messagebox.showerror("Erro", "Número de vagas deve ser um número inteiro.")
            return
        try:
            valor = float(valor_str.replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erro", "Valor mensal deve ser um número válido.")
            return

        vagas = int(vagas_str)

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM editais WHERE numero_edital = ?", (edital_selecionado,))
            resultado = cursor.fetchone()
            if not resultado:
                messagebox.showerror("Erro", "Edital não encontrado.")
                return
            edital_id = resultado[0]

            cursor.execute('''
                INSERT INTO modalidades (edital_id, nivel, vagas, valor_mensal)
                VALUES (?, ?, ?, ?)
            ''', (edital_id, nivel, vagas, valor))
            conn.commit()
            conn.close()

            messagebox.showinfo("Sucesso", "Modalidade cadastrada com sucesso!")
            self.limpar_campos_modalidade()

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao salvar modalidade:\n{str(e)}")

    def limpar_campos_modalidade(self):
        self.combo_editais_mod.set("")
        self.combo_nivel_mod.set("")
        self.entry_vagas.delete(0, tk.END)
        self.entry_valor.delete(0, tk.END)

    def criar_formulario_bolsistas(self):
        frame = ttk.Frame(self.tab_bolsistas)
        frame.pack(padx=20, pady=10)

        ttk.Label(frame, text="Edital:").grid(row=0, column=0, sticky='w', pady=4)
        self.combo_editais_bols = ttk.Combobox(frame, width=25, state="readonly")
        self.combo_editais_bols.grid(row=0, column=1, padx=5, pady=4)
        self.combo_editais_bols.bind("<<ComboboxSelected>>", self.atualizar_niveis_bolsista)

        ttk.Label(frame, text="Nível:").grid(row=0, column=2, sticky='w', pady=4, padx=(10,0))
        self.combo_nivel_bols = ttk.Combobox(frame, width=20, state="readonly")
        self.combo_nivel_bols.grid(row=0, column=3, padx=5, pady=4)

        ttk.Label(frame, text="Processo SEI:").grid(row=1, column=0, sticky='w', pady=4)
        self.entry_processo = ttk.Entry(frame, width=28)
        self.entry_processo.grid(row=1, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Nome do Bolsista:").grid(row=2, column=0, sticky='w', pady=4)
        self.entry_nome = ttk.Entry(frame, width=28)
        self.entry_nome.grid(row=2, column=1, padx=5, pady=4)

        ttk.Label(frame, text="CPF:").grid(row=2, column=2, sticky='w', pady=4, padx=(10,0))
        self.entry_cpf = ttk.Entry(frame, width=23)
        self.entry_cpf.grid(row=2, column=3, padx=5, pady=4)

        ttk.Label(frame, text="Orientador:").grid(row=3, column=0, sticky='w', pady=4)
        self.entry_orientador = ttk.Entry(frame, width=28)
        self.entry_orientador.grid(row=3, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Campus:").grid(row=4, column=0, sticky='w', pady=4)
        self.entry_campus = ttk.Entry(frame, width=28)
        self.entry_campus.grid(row=4, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Programa:").grid(row=5, column=0, sticky='w', pady=4)
        self.entry_programa = ttk.Entry(frame, width=28)
        self.entry_programa.grid(row=5, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Início no Curso (DD/MM/YYYY):").grid(row=6, column=0, sticky='w', pady=4)
        self.entry_data_inicio_curso = ttk.Entry(frame, width=28)
        self.entry_data_inicio_curso.grid(row=6, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Início da Bolsa (DD/MM/YYYY):").grid(row=6, column=2, sticky='w', pady=4, padx=(10,0))
        self.entry_data_inicio_bolsa = ttk.Entry(frame, width=23)
        self.entry_data_inicio_bolsa.grid(row=6, column=3, padx=5, pady=4)

        ttk.Label(frame, text="Meses de Duração:").grid(row=7, column=0, sticky='w', pady=4)
        self.entry_meses = ttk.Entry(frame, width=28)
        self.entry_meses.grid(row=7, column=1, padx=5, pady=4)

        ttk.Label(frame, text="Previsão Defesa (DD/MM/YYYY):").grid(row=8, column=0, sticky='w', pady=4)
        self.entry_defesa = ttk.Entry(frame, width=28)
        self.entry_defesa.grid(row=8, column=1, padx=5, pady=4)

        ttk.Label(frame, text="E-mail Bolsista:").grid(row=9, column=0, sticky='w', pady=4)
        self.entry_email_bols = ttk.Entry(frame, width=28)
        self.entry_email_bols.grid(row=9, column=1, padx=5, pady=4)

        ttk.Label(frame, text="E-mail Programa:").grid(row=10, column=0, sticky='w', pady=4)
        self.entry_email_prog = ttk.Entry(frame, width=28)
        self.entry_email_prog.grid(row=10, column=1, padx=5, pady=4)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=11, column=0, columnspan=4, pady=15)

        ttk.Button(btn_frame, text="Cadastrar", command=self.cadastrar_bolsista).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Limpar", command=self.limpar_campos_bolsista).pack(side='left', padx=5)

        self.carregar_editais_para_combo_bols()

    def carregar_editais_para_combo_bols(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, numero_edital FROM editais ORDER BY id")
            editais = cursor.fetchall()
            conn.close()

            if editais:
                self.combo_editais_bols['values'] = [f"{e[1]}" for e in editais]
                self.combo_editais_bols.set("")
            else:
                self.combo_editais_bols['values'] = ["Nenhum edital cadastrado"]
                self.combo_editais_bols.set("Nenhum edital cadastrado")
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar editais:\n{str(e)}")

    def atualizar_niveis_bolsista(self, event=None):
        edital = self.combo_editais_bols.get()
        if edital == "Nenhum edital cadastrado" or not edital:
            self.combo_nivel_bols['values'] = []
            self.combo_nivel_bols.set("")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT DISTINCT nivel FROM modalidades
                JOIN editais ON modalidades.edital_id = editais.id
                WHERE editais.numero_edital = ?
            ''', (edital,))
            niveis = [row[0] for row in cursor.fetchall()]
            conn.close()

            self.combo_nivel_bols['values'] = niveis
            self.combo_nivel_bols.set("")
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar níveis:\n{str(e)}")

    def cadastrar_bolsista(self):
        edital = self.combo_editais_bols.get()
        nivel = self.combo_nivel_bols.get()
        processo = self.entry_processo.get().strip()
        nome = self.entry_nome.get().strip()
        cpf = self.entry_cpf.get().strip()
        orientador = self.entry_orientador.get().strip()
        campus = self.entry_campus.get().strip()
        programa = self.entry_programa.get().strip()
        data_inicio_curso_br = self.entry_data_inicio_curso.get().strip()
        data_inicio_bolsa_br = self.entry_data_inicio_bolsa.get().strip()
        meses_str = self.entry_meses.get().strip()
        defesa_br = self.entry_defesa.get().strip()
        email_bols = self.entry_email_bols.get().strip()
        email_prog = self.entry_email_prog.get().strip()

        if edital == "Nenhum edital cadastrado" or not edital:
            messagebox.showerror("Erro", "Selecione um edital.")
            return
        if not nivel:
            messagebox.showerror("Erro", "Selecione o nível.")
            return
        if not nome:
            messagebox.showerror("Erro", "Nome do bolsista é obrigatório.")
            return
        if not data_inicio_curso_br:
            messagebox.showerror("Erro", "Data de início no curso é obrigatória.")
            return
        if not data_inicio_bolsa_br:
            messagebox.showerror("Erro", "Data de início da bolsa é obrigatória.")
            return
        if not meses_str.isdigit():
            messagebox.showerror("Erro", "Meses de duração deve ser um número inteiro.")
            return

        data_inicio_curso_iso = self.converter_data_br_para_iso(data_inicio_curso_br)
        data_inicio_bolsa_iso = self.converter_data_br_para_iso(data_inicio_bolsa_br)
        if not data_inicio_curso_iso:
            messagebox.showerror("Erro", "Data de início no curso inválida. Use DD/MM/YYYY.")
            return
        if not data_inicio_bolsa_iso:
            messagebox.showerror("Erro", "Data de início da bolsa inválida. Use DD/MM/YYYY.")
            return

        data_curso = datetime.strptime(data_inicio_curso_iso, "%Y-%m-%d")
        data_bolsa = datetime.strptime(data_inicio_bolsa_iso, "%Y-%m-%d")
        if data_bolsa < data_curso:
            messagebox.showerror("Erro", "A data de início da bolsa não pode ser anterior à data de início no curso.")
            return

        defesa_iso = self.converter_data_br_para_iso(defesa_br) if defesa_br else None

        meses = int(meses_str)
        data_fim_iso = self.calcular_data_fim_bolsa(data_inicio_bolsa_iso, meses)
        if not data_fim_iso:
            messagebox.showerror("Erro", "Erro ao calcular data final da bolsa.")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM editais WHERE numero_edital = ?", (edital,))
            resultado = cursor.fetchone()
            if not resultado:
                messagebox.showerror("Erro", "Edital não encontrado.")
                return
            edital_id = resultado[0]

            cursor.execute('''
                INSERT INTO bolsistas (
                    edital_id, processo_sei, nome, cpf, orientador, campus, programa,
                    nivel, data_inicio_curso, data_inicio_bolsa, meses_duracao, data_fim_bolsa,
                    previsao_defesa, email_bolsista, email_programa, status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                edital_id, processo, nome, cpf, orientador, campus, programa,
                nivel, data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
                defesa_iso, email_bols, email_prog, 'ativo'
            ))
            conn.commit()
            conn.close()

            messagebox.showinfo("Sucesso", "Bolsista cadastrado com sucesso!")
            self.limpar_campos_bolsista()

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao salvar bolsista:\n{str(e)}")

    def limpar_campos_bolsista(self):
        self.combo_editais_bols.set("")
        self.combo_nivel_bols.set("")
        self.entry_processo.delete(0, tk.END)
        self.entry_nome.delete(0, tk.END)
        self.entry_cpf.delete(0, tk.END)
        self.entry_orientador.delete(0, tk.END)
        self.entry_campus.delete(0, tk.END)
        self.entry_programa.delete(0, tk.END)
        self.entry_data_inicio_curso.delete(0, tk.END)
        self.entry_data_inicio_bolsa.delete(0, tk.END)
        self.entry_meses.delete(0, tk.END)
        self.entry_defesa.delete(0, tk.END)
        self.entry_email_bols.delete(0, tk.END)
        self.entry_email_prog.delete(0, tk.END)

    def criar_aba_consulta(self):
        frame_top = ttk.Frame(self.tab_consulta)
        frame_top.pack(fill='x', padx=10, pady=5)

        ttk.Label(frame_top, text="Edital:").grid(row=0, column=0, sticky='w', padx=(0,5))
        self.combo_consulta_editais = ttk.Combobox(frame_top, width=25, state="readonly")
        self.combo_consulta_editais.grid(row=0, column=1, padx=5, pady=2)
        self.combo_consulta_editais.bind("<<ComboboxSelected>>", self.carregar_dados_consulta)

        ttk.Label(frame_top, text="Projeto:").grid(row=0, column=2, sticky='w', padx=(15,5))
        self.entry_projeto_consulta = ttk.Entry(frame_top, width=20)
        self.entry_projeto_consulta.grid(row=0, column=3, padx=5, pady=2)
        ttk.Button(frame_top, text="Buscar", command=self.buscar_por_projeto).grid(row=0, column=4, padx=5, pady=2)

        btn_export = ttk.Menubutton(frame_top, text="Exportar Relatório", width=18)
        menu = tk.Menu(btn_export, tearoff=0)
        menu.add_command(label="TXT", command=lambda: self.exportar_relatorio("txt"))
        menu.add_command(label="Excel", command=lambda: self.exportar_relatorio("excel"))
        menu.add_command(label="PDF", command=lambda: self.exportar_relatorio("pdf"))
        btn_export["menu"] = menu
        btn_export.grid(row=0, column=5, padx=(20,0), pady=2, sticky='e')

        frame_top.columnconfigure(1, weight=1)
        frame_top.columnconfigure(3, weight=1)
        frame_top.columnconfigure(5, weight=0)

        self.notebook_consulta = ttk.Notebook(self.tab_consulta)
        self.notebook_consulta.pack(fill='both', expand=True, padx=10, pady=(0,10))

        self.tab_consulta_editais = ttk.Frame(self.notebook_consulta)
        self.tab_consulta_modalidades = ttk.Frame(self.notebook_consulta)
        self.tab_consulta_bolsistas = ttk.Frame(self.notebook_consulta)
        self.notebook_consulta.add(self.tab_consulta_editais, text="Editais")
        self.notebook_consulta.add(self.tab_consulta_modalidades, text="Modalidades")
        self.notebook_consulta.add(self.tab_consulta_bolsistas, text="Bolsistas Ativos")

        # Treeview para Editais
        cols_editais = ("ID", "Número", "Descrição", "Agência", "Projeto", "Descrição Projeto")
        self.tree_editais = ttk.Treeview(self.tab_consulta_editais, columns=cols_editais, show='headings')
        for col in cols_editais:
            self.tree_editais.heading(col, text=col if col != "Descrição Projeto" else "Desc. Projeto")
            if col == "ID": self.tree_editais.column(col, width=50)
            elif col == "Número": self.tree_editais.column(col, width=100)
            elif col == "Descrição": self.tree_editais.column(col, width=150)
            elif col == "Agência": self.tree_editais.column(col, width=120)
            elif col == "Projeto": self.tree_editais.column(col, width=100)
            else: self.tree_editais.column(col, width=150)
        scrollbar_editais = ttk.Scrollbar(self.tab_consulta_editais, orient="vertical", command=self.tree_editais.yview)
        self.tree_editais.configure(yscroll=scrollbar_editais.set)
        self.tree_editais.pack(side='left', fill='both', expand=True)
        scrollbar_editais.pack(side='right', fill='y')
        self.tree_editais.bind("<Double-1>", self.editar_edital_selecionado)
        self.tree_editais.bind("<Button-3>", self.excluir_edital_selecionado)  # Clique direito

        # Treeview para Modalidades
        cols_mod = ("ID", "Edital", "Nível", "Vagas", "Valor")
        self.tree_modalidades = ttk.Treeview(self.tab_consulta_modalidades, columns=cols_mod, show='headings')
        for col in cols_mod:
            self.tree_modalidades.heading(col, text=col)
            if col == "ID": self.tree_modalidades.column(col, width=50)
            elif col == "Edital": self.tree_modalidades.column(col, width=120)
            elif col == "Nível": self.tree_modalidades.column(col, width=120)
            elif col == "Vagas": self.tree_modalidades.column(col, width=80)
            else: self.tree_modalidades.column(col, width=120)
        scrollbar_mod = ttk.Scrollbar(self.tab_consulta_modalidades, orient="vertical", command=self.tree_modalidades.yview)
        self.tree_modalidades.configure(yscroll=scrollbar_mod.set)
        self.tree_modalidades.pack(side='left', fill='both', expand=True)
        scrollbar_mod.pack(side='right', fill='y')
        self.tree_modalidades.bind("<Double-1>", self.editar_modalidade_selecionada)
        self.tree_modalidades.bind("<Button-3>", self.excluir_modalidade_selecionada)  # Clique direito

        # Treeview para Bolsistas
        cols_bols = ("ID", "Nome", "CPF", "Nível", "Início Bolsa", "Fim Bolsa", "Status")
        self.tree_bolsistas = ttk.Treeview(self.tab_consulta_bolsistas, columns=cols_bols, show='headings')
        for col in cols_bols:
            self.tree_bolsistas.heading(col, text=col)
            if col == "ID": self.tree_bolsistas.column(col, width=50)
            elif col == "Nome": self.tree_bolsistas.column(col, width=180)
            elif col == "CPF": self.tree_bolsistas.column(col, width=120)
            else: self.tree_bolsistas.column(col, width=100)
        scrollbar_bols = ttk.Scrollbar(self.tab_consulta_bolsistas, orient="vertical", command=self.tree_bolsistas.yview)
        self.tree_bolsistas.configure(yscroll=scrollbar_bols.set)
        self.tree_bolsistas.pack(side='left', fill='both', expand=True)
        scrollbar_bols.pack(side='right', fill='y')
        self.tree_bolsistas.bind("<Double-1>", self.editar_bolsista_selecionado)
        self.tree_bolsistas.bind("<Button-3>", self.excluir_bolsista_selecionado)  # Clique direito

        self.carregar_editais_consulta()

    # === EXCLUSÃO DE EDITAIS ===
    def excluir_edital_selecionado(self, event):
        item_id = self.tree_editais.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_editais.item(item_id)
        edital_id = item['values'][0]
        numero_edital = item['values'][1]

        # Verificar dependências
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM modalidades WHERE edital_id = ?", (edital_id,))
        modalidades = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM bolsistas WHERE edital_id = ?", (edital_id,))
        bolsistas = cursor.fetchone()[0]
        conn.close()

        if modalidades > 0 or bolsistas > 0:
            messagebox.showwarning("Aviso", 
                f"O edital '{numero_edital}' não pode ser excluído porque tem:\n"
                f"- {modalidades} modalidade(s)\n"
                f"- {bolsistas} bolsista(s)\n\n"
                "Exclua as dependências primeiro.")
            return

        if messagebox.askyesno("Confirmar Exclusão", 
                              f"Tem certeza que deseja excluir o edital '{numero_edital}'?\n"
                              "Esta ação não pode ser desfeita."):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM editais WHERE id = ?", (edital_id,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Sucesso", "Edital excluído com sucesso!")
                self.carregar_dados_consulta()
            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao excluir edital:\n{str(e)}")

    # === EXCLUSÃO DE MODALIDADES ===
    def excluir_modalidade_selecionada(self, event):
        item_id = self.tree_modalidades.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_modalidades.item(item_id)
        modalidade_id = item['values'][0]
        edital_num = item['values'][1]
        nivel = item['values'][2]

        # Verificar dependências
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM bolsistas 
            JOIN editais ON bolsistas.edital_id = editais.id
            WHERE editais.numero_edital = ? AND bolsistas.nivel = ?
        """, (edital_num, nivel))
        bolsistas = cursor.fetchone()[0]
        conn.close()

        if bolsistas > 0:
            messagebox.showwarning("Aviso", 
                f"A modalidade '{nivel}' do edital '{edital_num}' não pode ser excluída\n"
                f"porque tem {bolsistas} bolsista(s) vinculado(s).\n\n"
                "Exclua os bolsistas primeiro ou altere o nível deles.")
            return

        if messagebox.askyesno("Confirmar Exclusão", 
                              f"Tem certeza que deseja excluir a modalidade '{nivel}'\n"
                              f"do edital '{edital_num}'?\n"
                              "Esta ação não pode ser desfeita."):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM modalidades WHERE id = ?", (modalidade_id,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Sucesso", "Modalidade excluída com sucesso!")
                self.carregar_dados_consulta()
            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao excluir modalidade:\n{str(e)}")

    # === EXCLUSÃO DE BOLSISTAS ===
    def excluir_bolsista_selecionado(self, event):
        item_id = self.tree_bolsistas.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_bolsistas.item(item_id)
        bolsista_id = item['values'][0]
        nome = item['values'][1]

        if messagebox.askyesno("Confirmar Exclusão", 
                              f"Tem certeza que deseja excluir o bolsista '{nome}'?\n"
                              "Esta ação não pode ser desfeita."):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM bolsistas WHERE id = ?", (bolsista_id,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Sucesso", "Bolsista excluído com sucesso!")
                self.carregar_dados_consulta()
            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Erro ao excluir bolsista:\n{str(e)}")

    def carregar_editais_consulta(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, numero_edital FROM editais ORDER BY id")
            editais = cursor.fetchall()
            conn.close()
            self.combo_consulta_editais['values'] = ["Todos"] + [e[1] for e in editais]
            self.combo_consulta_editais.set("Todos")
            self.carregar_dados_consulta()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar editais para consulta:\n{str(e)}")

    def carregar_dados_consulta(self, event=None):
        for item in self.tree_editais.get_children():
            self.tree_editais.delete(item)
        for item in self.tree_modalidades.get_children():
            self.tree_modalidades.delete(item)
        for item in self.tree_bolsistas.get_children():
            self.tree_bolsistas.delete(item)

        edital_selecionado = self.combo_consulta_editais.get()
        if edital_selecionado == "Todos":
            self.carregar_todos_dados()
        elif edital_selecionado:
            self.carregar_dados_por_edital(edital_selecionado)

    def carregar_todos_dados(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            # Editais
            cursor.execute('''
                SELECT id, numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto
                FROM editais
                ORDER BY id
            ''')
            for row in cursor.fetchall():
                self.tree_editais.insert("", "end", values=row)

            # Modalidades
            cursor.execute('''
                SELECT modalidades.id, editais.numero_edital, nivel, vagas, valor_mensal
                FROM modalidades
                JOIN editais ON modalidades.edital_id = editais.id
                ORDER BY editais.numero_edital, nivel
            ''')
            for row in cursor.fetchall():
                self.tree_modalidades.insert("", "end", values=(row[0], row[1], row[2], row[3], f"R$ {row[4]:.2f}"))

            # Bolsistas
            cursor.execute('''
                SELECT bolsistas.id, nome, cpf, nivel, data_inicio_bolsa, data_fim_bolsa, status
                FROM bolsistas
                WHERE status = 'ativo'
                ORDER BY nome
            ''')
            for row in cursor.fetchall():
                ini = self.converter_data_iso_para_br(row[4])
                fim = self.converter_data_iso_para_br(row[5])
                self.tree_bolsistas.insert("", "end", values=(row[0], row[1], row[2], row[3], ini, fim, row[6]))
            conn.close()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar todos os dados:\n{str(e)}")

    def carregar_dados_por_edital(self, numero_edital):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            # Editais
            cursor.execute('''
                SELECT id, numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto
                FROM editais
                WHERE numero_edital = ?
            ''', (numero_edital,))
            for row in cursor.fetchall():
                self.tree_editais.insert("", "end", values=row)

            # Modalidades
            cursor.execute('''
                SELECT modalidades.id, editais.numero_edital, nivel, vagas, valor_mensal
                FROM modalidades
                JOIN editais ON modalidades.edital_id = editais.id
                WHERE editais.numero_edital = ?
                ORDER BY nivel
            ''', (numero_edital,))
            for row in cursor.fetchall():
                self.tree_modalidades.insert("", "end", values=(row[0], row[1], row[2], row[3], f"R$ {row[4]:.2f}"))

            # Bolsistas
            cursor.execute('''
                SELECT bolsistas.id, nome, cpf, nivel, data_inicio_bolsa, data_fim_bolsa, status
                FROM bolsistas
                JOIN editais ON bolsistas.edital_id = editais.id
                WHERE editais.numero_edital = ? AND bolsistas.status = 'ativo'
                ORDER BY nome
            ''', (numero_edital,))
            for row in cursor.fetchall():
                ini = self.converter_data_iso_para_br(row[4])
                fim = self.converter_data_iso_para_br(row[5])
                self.tree_bolsistas.insert("", "end", values=(row[0], row[1], row[2], row[3], ini, fim, row[6]))
            conn.close()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do edital:\n{str(e)}")

    def buscar_por_projeto(self):
        codigo_projeto = self.entry_projeto_consulta.get().strip()
        if not codigo_projeto:
            messagebox.showwarning("Aviso", "Informe o código do projeto.")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id, numero_edital FROM editais WHERE codigo_projeto LIKE ?", (f"%{codigo_projeto}%",))
            editais = cursor.fetchall()
            if not editais:
                messagebox.showinfo("Info", "Nenhum edital encontrado com esse código de projeto.")
                return

            for item in self.tree_editais.get_children():
                self.tree_editais.delete(item)
            for item in self.tree_modalidades.get_children():
                self.tree_modalidades.delete(item)
            for item in self.tree_bolsistas.get_children():
                self.tree_bolsistas.delete(item)

            for edital_id, numero in editais:
                # Editais
                cursor.execute('''
                    SELECT id, numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto
                    FROM editais WHERE id = ?
                ''', (edital_id,))
                for row in cursor.fetchall():
                    self.tree_editais.insert("", "end", values=row)

                # Modalidades
                cursor.execute('''
                    SELECT modalidades.id, ?, nivel, vagas, valor_mensal
                    FROM modalidades WHERE edital_id = ?
                ''', (numero, edital_id))
                for row in cursor.fetchall():
                    self.tree_modalidades.insert("", "end", values=(row[0], row[1], row[2], row[3], f"R$ {row[4]:.2f}"))

                # Bolsistas
                cursor.execute('''
                    SELECT bolsistas.id, nome, cpf, nivel, data_inicio_bolsa, data_fim_bolsa, status
                    FROM bolsistas WHERE edital_id = ? AND status = 'ativo'
                    ORDER BY nome
                ''', (edital_id,))
                for row in cursor.fetchall():
                    ini = self.converter_data_iso_para_br(row[4])
                    fim = self.converter_data_iso_para_br(row[5])
                    self.tree_bolsistas.insert("", "end", values=(row[0], row[1], row[2], row[3], ini, fim, row[6]))
            conn.close()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro na busca por projeto:\n{str(e)}")

    # === EDIÇÃO (mesmo que antes) ===
    def editar_edital_selecionado(self, event):
        item_id = self.tree_editais.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_editais.item(item_id)
        edital_id = item['values'][0]
        self.abrir_janela_edicao_edital(edital_id)

    def abrir_janela_edicao_edital(self, edital_id):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto
                FROM editais WHERE id = ?
            ''', (edital_id,))
            dados = cursor.fetchone()
            conn.close()
            if not dados:
                messagebox.showerror("Erro", "Edital não encontrado.")
                return
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar edital:\n{str(e)}")
            return

        janela = tk.Toplevel(self.root)
        janela.title(f"Editar Edital (ID: {edital_id})")
        janela.geometry("600x400")

        frame = ttk.Frame(janela, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="Número do Edital:").grid(row=0, column=0, sticky='w', pady=5)
        entry_numero = ttk.Entry(frame, width=40)
        entry_numero.grid(row=0, column=1, pady=5)
        entry_numero.insert(0, dados[0])

        ttk.Label(frame, text="Descrição:").grid(row=1, column=0, sticky='w', pady=5)
        entry_desc = ttk.Entry(frame, width=40)
        entry_desc.grid(row=1, column=1, pady=5)
        entry_desc.insert(0, dados[1] or "")

        ttk.Label(frame, text="Agência de Fomento:").grid(row=2, column=0, sticky='w', pady=5)
        entry_agencia = ttk.Entry(frame, width=40)
        entry_agencia.grid(row=2, column=1, pady=5)
        entry_agencia.insert(0, dados[2])

        ttk.Label(frame, text="Código do Projeto:").grid(row=3, column=0, sticky='w', pady=5)
        entry_projeto = ttk.Entry(frame, width=40)
        entry_projeto.grid(row=3, column=1, pady=5)
        entry_projeto.insert(0, dados[3] or "")

        ttk.Label(frame, text="Descrição do Projeto:").grid(row=4, column=0, sticky='w', pady=5)
        entry_desc_proj = ttk.Entry(frame, width=40)
        entry_desc_proj.grid(row=4, column=1, pady=5)
        entry_desc_proj.insert(0, dados[4] or "")

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text="Salvar", command=lambda: self.salvar_edicao_edital(
            edital_id, entry_numero.get().strip(), entry_desc.get().strip(),
            entry_agencia.get().strip(), entry_projeto.get().strip(), entry_desc_proj.get().strip(), janela
        )).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

    def salvar_edicao_edital(self, edital_id, numero, descricao, agencia, codigo_projeto, descricao_projeto, janela):
        if not numero:
            messagebox.showerror("Erro", "Número do edital é obrigatório.", parent=janela)
            return
        if not agencia:
            messagebox.showerror("Erro", "Agência de fomento é obrigatória.", parent=janela)
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE editais SET
                    numero_edital = ?, descricao = ?, agencia_fomento = ?,
                    codigo_projeto = ?, descricao_projeto = ?
                WHERE id = ?
            ''', (numero, descricao, agencia, codigo_projeto or None, descricao_projeto or None, edital_id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Sucesso", "Edital atualizado com sucesso!", parent=janela)
            janela.destroy()
            self.carregar_dados_consulta()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao atualizar edital:\n{str(e)}", parent=janela)

    def editar_modalidade_selecionada(self, event):
        item_id = self.tree_modalidades.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_modalidades.item(item_id)
        modalidade_id = item['values'][0]
        self.abrir_janela_edicao_modalidade(modalidade_id)

    def abrir_janela_edicao_modalidade(self, modalidade_id):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT edital_id, nivel, vagas, valor_mensal
                FROM modalidades WHERE id = ?
            ''', (modalidade_id,))
            dados = cursor.fetchone()
            if dados:
                cursor.execute("SELECT numero_edital FROM editais WHERE id = ?", (dados[0],))
                edital_num = cursor.fetchone()[0]
                dados = (edital_num, dados[1], dados[2], dados[3])
            conn.close()
            if not dados:
                messagebox.showerror("Erro", "Modalidade não encontrada.")
                return
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar modalidade:\n{str(e)}")
            return

        janela = tk.Toplevel(self.root)
        janela.title(f"Editar Modalidade (ID: {modalidade_id})")
        janela.geometry("500x300")

        frame = ttk.Frame(janela, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="Edital:").grid(row=0, column=0, sticky='w', pady=5)
        entry_edital = ttk.Entry(frame, width=30)
        entry_edital.grid(row=0, column=1, pady=5)
        entry_edital.insert(0, dados[0])
        entry_edital.config(state='readonly')

        ttk.Label(frame, text="Nível:").grid(row=1, column=0, sticky='w', pady=5)
        entry_nivel = ttk.Entry(frame, width=30)
        entry_nivel.grid(row=1, column=1, pady=5)
        entry_nivel.insert(0, dados[1])
        entry_nivel.config(state='readonly')

        ttk.Label(frame, text="Vagas:").grid(row=2, column=0, sticky='w', pady=5)
        entry_vagas = ttk.Entry(frame, width=30)
        entry_vagas.grid(row=2, column=1, pady=5)
        entry_vagas.insert(0, str(dados[2]))

        ttk.Label(frame, text="Valor Mensal (R$):").grid(row=3, column=0, sticky='w', pady=5)
        entry_valor = ttk.Entry(frame, width=30)
        entry_valor.grid(row=3, column=1, pady=5)
        entry_valor.insert(0, str(dados[3]))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text="Salvar", command=lambda: self.salvar_edicao_modalidade(
            modalidade_id, entry_vagas.get().strip(), entry_valor.get().strip(), janela
        )).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

    def salvar_edicao_modalidade(self, modalidade_id, vagas_str, valor_str, janela):
        if not vagas_str.isdigit():
            messagebox.showerror("Erro", "Vagas deve ser um número inteiro.", parent=janela)
            return
        try:
            valor = float(valor_str.replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erro", "Valor deve ser um número válido.", parent=janela)
            return

        vagas = int(vagas_str)
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE modalidades SET vagas = ?, valor_mensal = ?
                WHERE id = ?
            ''', (vagas, valor, modalidade_id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Sucesso", "Modalidade atualizada com sucesso!", parent=janela)
            janela.destroy()
            self.carregar_dados_consulta()
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao atualizar modalidade:\n{str(e)}", parent=janela)

    def editar_bolsista_selecionado(self, event):
        item_id = self.tree_bolsistas.identify_row(event.y)
        if not item_id:
            return
        item = self.tree_bolsistas.item(item_id)
        bolsista_id = item['values'][0]
        self.abrir_janela_edicao_bolsista(bolsista_id)

    def abrir_janela_edicao_bolsista(self, bolsista_id):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT
                    bolsistas.edital_id, numero_edital, processo_sei, nome, cpf, orientador,
                    campus, programa, nivel, data_inicio_curso, data_inicio_bolsa,
                    meses_duracao, previsao_defesa, email_bolsista, email_programa, status
                FROM bolsistas
                JOIN editais ON bolsistas.edital_id = editais.id
                WHERE bolsistas.id = ?
            ''', (bolsista_id,))
            dados = cursor.fetchone()
            conn.close()
            if not dados:
                messagebox.showerror("Erro", "Bolsista não encontrado.")
                return
        except sqlite3.Error as e:
            messagebox.showerror("Erro", f"Erro ao carregar bolsista:\n{str(e)}")
            return

        janela = tk.Toplevel(self.root)
        janela.title(f"Editar Bolsista (ID: {bolsista_id})")
        janela.geometry("700x600")
        janela.resizable(False, False)

        canvas = tk.Canvas(janela, bg='#2d2d2d')
        scrollbar = ttk.Scrollbar(janela, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set, highlightthickness=0)

        def on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", on_canvas_configure)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        style = ttk.Style()
        style.configure("TLabel", background='#2d2d2d', foreground='white')
        style.configure("TEntry", fieldbackground='#3c3c3c', foreground='white')
        style.configure("TCombobox", fieldbackground='#3c3c3c', foreground='white')
        style.configure("TFrame", background='#2d2d2d')

        content_frame = ttk.Frame(scrollable_frame)
        content_frame.pack(padx=20, pady=20)

        ttk.Label(content_frame, text="Edital:").grid(row=0, column=0, sticky='w', pady=4)
        entry_edital = ttk.Entry(content_frame, width=30)
        entry_edital.grid(row=0, column=1, padx=5, pady=4)
        entry_edital.insert(0, dados[1])
        entry_edital.config(state='readonly')

        ttk.Label(content_frame, text="Processo SEI:").grid(row=1, column=0, sticky='w', pady=4)
        entry_processo = ttk.Entry(content_frame, width=30)
        entry_processo.grid(row=1, column=1, padx=5, pady=4)
        entry_processo.insert(0, dados[2] or "")

        ttk.Label(content_frame, text="Nome:").grid(row=2, column=0, sticky='w', pady=4)
        entry_nome = ttk.Entry(content_frame, width=30)
        entry_nome.grid(row=2, column=1, padx=5, pady=4)
        entry_nome.insert(0, dados[3])

        ttk.Label(content_frame, text="CPF:").grid(row=3, column=0, sticky='w', pady=4)
        entry_cpf = ttk.Entry(content_frame, width=30)
        entry_cpf.grid(row=3, column=1, padx=5, pady=4)
        entry_cpf.insert(0, dados[4] or "")

        ttk.Label(content_frame, text="Orientador:").grid(row=4, column=0, sticky='w', pady=4)
        entry_orientador = ttk.Entry(content_frame, width=30)
        entry_orientador.grid(row=4, column=1, padx=5, pady=4)
        entry_orientador.insert(0, dados[5] or "")

        ttk.Label(content_frame, text="Campus:").grid(row=5, column=0, sticky='w', pady=4)
        entry_campus = ttk.Entry(content_frame, width=30)
        entry_campus.grid(row=5, column=1, padx=5, pady=4)
        entry_campus.insert(0, dados[6] or "")

        ttk.Label(content_frame, text="Programa:").grid(row=6, column=0, sticky='w', pady=4)
        entry_programa = ttk.Entry(content_frame, width=30)
        entry_programa.grid(row=6, column=1, padx=5, pady=4)
        entry_programa.insert(0, dados[7] or "")

        ttk.Label(content_frame, text="Nível:").grid(row=7, column=0, sticky='w', pady=4)
        entry_nivel = ttk.Entry(content_frame, width=30)
        entry_nivel.grid(row=7, column=1, padx=5, pady=4)
        entry_nivel.insert(0, dados[8])
        entry_nivel.config(state='readonly')

        ttk.Label(content_frame, text="Início no Curso (DD/MM/YYYY):").grid(row=8, column=0, sticky='w', pady=4)
        entry_inicio_curso = ttk.Entry(content_frame, width=30)
        entry_inicio_curso.grid(row=8, column=1, padx=5, pady=4)
        entry_inicio_curso.insert(0, self.converter_data_iso_para_br(dados[9]))

        ttk.Label(content_frame, text="Início da Bolsa (DD/MM/YYYY):").grid(row=9, column=0, sticky='w', pady=4)
        entry_inicio_bolsa = ttk.Entry(content_frame, width=30)
        entry_inicio_bolsa.grid(row=9, column=1, padx=5, pady=4)
        entry_inicio_bolsa.insert(0, self.converter_data_iso_para_br(dados[10]))

        ttk.Label(content_frame, text="Meses de Duração:").grid(row=10, column=0, sticky='w', pady=4)
        entry_meses = ttk.Entry(content_frame, width=30)
        entry_meses.grid(row=10, column=1, padx=5, pady=4)
        entry_meses.insert(0, str(dados[11]))

        ttk.Label(content_frame, text="Previsão Defesa (DD/MM/YYYY):").grid(row=11, column=0, sticky='w', pady=4)
        entry_defesa = ttk.Entry(content_frame, width=30)
        entry_defesa.grid(row=11, column=1, padx=5, pady=4)
        entry_defesa.insert(0, self.converter_data_iso_para_br(dados[12]) if dados[12] else "")

        ttk.Label(content_frame, text="E-mail Bolsista:").grid(row=12, column=0, sticky='w', pady=4)
        entry_email_bols = ttk.Entry(content_frame, width=30)
        entry_email_bols.grid(row=12, column=1, padx=5, pady=4)
        entry_email_bols.insert(0, dados[13] or "")

        ttk.Label(content_frame, text="E-mail Programa:").grid(row=13, column=0, sticky='w', pady=4)
        entry_email_prog = ttk.Entry(content_frame, width=30)
        entry_email_prog.grid(row=13, column=1, padx=5, pady=4)
        entry_email_prog.insert(0, dados[14] or "")

        ttk.Label(content_frame, text="Status:").grid(row=14, column=0, sticky='w', pady=4)
        combo_status = ttk.Combobox(content_frame, width=28, state="readonly")
        combo_status['values'] = ("ativo", "desligado", "substituido")
        combo_status.grid(row=14, column=1, padx=5, pady=4)
        combo_status.set(dados[15])

        btn_frame = ttk.Frame(content_frame)
        btn_frame.grid(row=15, column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text="Salvar", command=lambda: self.salvar_edicao_bolsista(
            bolsista_id, entry_processo.get().strip(), entry_nome.get().strip(),
            entry_cpf.get().strip(), entry_orientador.get().strip(),
            entry_campus.get().strip(), entry_programa.get().strip(),
            entry_inicio_curso.get().strip(), entry_inicio_bolsa.get().strip(),
            entry_meses.get().strip(), entry_defesa.get().strip(),
            entry_email_bols.get().strip(), entry_email_prog.get().strip(),
            combo_status.get(), janela
        )).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

    def salvar_edicao_bolsista(self, bolsista_id, processo, nome, cpf, orientador, campus, programa,
                              data_inicio_curso_br, data_inicio_bolsa_br, meses_str, defesa_br,
                              email_bols, email_prog, status, janela):
        if not nome:
            messagebox.showerror("Erro", "Nome é obrigatório.", parent=janela)
            return
        if not data_inicio_curso_br:
            messagebox.showerror("Erro", "Data de início no curso é obrigatória.", parent=janela)
            return
        if not data_inicio_bolsa_br:
            messagebox.showerror("Erro", "Data de início da bolsa é obrigatória.", parent=janela)
            return
        if not meses_str.isdigit():
            messagebox.showerror("Erro", "Meses de duração deve ser um número inteiro.", parent=janela)
            return

        data_inicio_curso_iso = self.converter_data_br_para_iso(data_inicio_curso_br)
        data_inicio_bolsa_iso = self.converter_data_br_para_iso(data_inicio_bolsa_br)
        if not data_inicio_curso_iso:
            messagebox.showerror("Erro", "Data de início no curso inválida. Use DD/MM/YYYY.", parent=janela)
            return
        if not data_inicio_bolsa_iso:
            messagebox.showerror("Erro", "Data de início da bolsa inválida. Use DD/MM/YYYY.", parent=janela)
            return

        data_curso = datetime.strptime(data_inicio_curso_iso, "%Y-%m-%d")
        data_bolsa = datetime.strptime(data_inicio_bolsa_iso, "%Y-%m-%d")
        if data_bolsa < data_curso:
            messagebox.showerror("Erro", "Início da bolsa não pode ser antes do início no curso.", parent=janela)
            return

        defesa_iso = self.converter_data_br_para_iso(defesa_br) if defesa_br else None
        meses = int(meses_str)
        data_fim_iso = self.calcular_data_fim_bolsa(data_inicio_bolsa_iso, meses)
        if not data_fim_iso:
            messagebox.showerror("Erro", "Erro ao calcular data final da bolsa.", parent=janela)
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE bolsistas SET
                    processo_sei = ?, nome = ?, cpf = ?, orientador = ?, campus = ?, programa = ?,
                    data_inicio_curso = ?, data_inicio_bolsa = ?, meses_duracao = ?, data_fim_bolsa = ?,
                    previsao_defesa = ?, email_bolsista = ?, email_programa = ?, status = ?
                WHERE id = ?
            ''', (
                processo, nome, cpf, orientador, campus, programa,
                data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
                defesa_iso, email_bols, email_prog, status, bolsista_id
            ))
            conn.commit()
            conn.close()

            messagebox.showinfo("Sucesso", "Bolsista atualizado com sucesso!", parent=janela)
            janela.destroy()
            self.carregar_dados_consulta()

        except sqlite3.Error as e:
            messagebox.showerror("Erro de Banco de Dados", f"Erro ao atualizar bolsista:\n{str(e)}", parent=janela)

    # === RELATÓRIOS (mesmo que antes) ===
    def exportar_relatorio(self, formato):
        edital_selecionado = self.combo_consulta_editais.get()
        if edital_selecionado == "Todos":
            dados_bolsistas = self.obter_dados_bolsistas_todos()
            dados_modalidades = self.obter_dados_modalidades_todos()
            nome_arquivo_base = "relatorio_todos_editais"
        else:
            dados_bolsistas = self.obter_dados_bolsistas_por_edital(edital_selecionado)
            dados_modalidades = self.obter_dados_modalidades_por_edital(edital_selecionado)
            nome_arquivo_base = f"relatorio_{edital_selecionado.replace('/', '_')}"

        if not dados_bolsistas and not dados_modalidades:
            messagebox.showinfo("Info", "Nenhum dado para exportar.")
            return

        pasta = filedialog.askdirectory(title="Escolha a pasta para salvar o relatório")
        if not pasta:
            return

        caminho = os.path.join(pasta, nome_arquivo_base)

        try:
            if formato == "txt":
                self.gerar_txt(caminho + ".txt", dados_modalidades, dados_bolsistas)
            elif formato == "excel":
                self.gerar_excel(caminho + ".xlsx", dados_modalidades, dados_bolsistas)
            elif formato == "pdf":
                self.gerar_pdf(caminho + ".pdf", dados_modalidades, dados_bolsistas, edital_selecionado)
            messagebox.showinfo("Sucesso", f"Relatório {formato.upper()} salvo com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{str(e)}")

    def obter_dados_bolsistas_todos(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nome, cpf, nivel, data_inicio_bolsa, data_fim_bolsa, status
            FROM bolsistas
            WHERE status = 'ativo'
            ORDER BY nome
        ''')
        dados = cursor.fetchall()
        conn.close()
        return dados

    def obter_dados_bolsistas_por_edital(self, numero_edital):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nome, cpf, nivel, data_inicio_bolsa, data_fim_bolsa, status
            FROM bolsistas
            JOIN editais ON bolsistas.edital_id = editais.id
            WHERE editais.numero_edital = ? AND bolsistas.status = 'ativo'
            ORDER BY nome
        ''', (numero_edital,))
        dados = cursor.fetchall()
        conn.close()
        return dados

    def obter_dados_modalidades_todos(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nivel, vagas, valor_mensal
            FROM modalidades
            JOIN editais ON modalidades.edital_id = editais.id
            ORDER BY editais.numero_edital, nivel
        ''')
        dados = cursor.fetchall()
        conn.close()
        return dados

    def obter_dados_modalidades_por_edital(self, numero_edital):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nivel, vagas, valor_mensal
            FROM modalidades
            JOIN editais ON modalidades.edital_id = editais.id
            WHERE editais.numero_edital = ?
            ORDER BY nivel
        ''', (numero_edital,))
        dados = cursor.fetchall()
        conn.close()
        return dados

    def gerar_txt(self, caminho, modalidades, bolsistas):
        with open(caminho, 'w', encoding='utf-8') as f:
            f.write("RELATÓRIO DE GESTÃO DE EDITAIS\n")
            f.write("=" * 50 + "\n\n")

            f.write("MODALIDADES:\n")
            f.write("-" * 30 + "\n")
            for nivel, vagas, valor in modalidades:
                f.write(f"Nível: {nivel}\n")
                f.write(f"Vagas: {vagas}\n")
                f.write(f"Valor: R$ {valor:.2f}\n\n")

            f.write("BOLSISTAS ATIVOS:\n")
            f.write("-" * 30 + "\n")
            for nome, cpf, nivel, ini, fim, status in bolsistas:
                ini_br = self.converter_data_iso_para_br(ini)
                fim_br = self.converter_data_iso_para_br(fim)
                f.write(f"Nome: {nome}\n")
                f.write(f"CPF: {cpf}\n")
                f.write(f"Nível: {nivel}\n")
                f.write(f"Início Bolsa: {ini_br}\n")
                f.write(f"Fim Bolsa: {fim_br}\n")
                f.write(f"Status: {status}\n\n")

    def gerar_excel(self, caminho, modalidades, bolsistas):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Editais"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="363636", end_color="363636", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A1:F1')
        ws['A1'] = "RELATÓRIO DE GESTÃO DE EDITAIS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        ws.append([])

        ws.append(["MODALIDADES"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.row_dimensions[ws.max_row].height = 25

        if modalidades:
            ws.append(["Nível", "Vagas", "Valor Mensal (R$)"])
            header_row = ws.max_row
            for col in range(1, 4):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for nivel, vagas, valor in modalidades:
                ws.append([nivel, vagas, f"R$ {valor:.2f}"])
                for col in range(1, 4):
                    ws.cell(row=ws.max_row, column=col).border = thin_border
        else:
            ws.append(["Nenhuma modalidade encontrada."])

        ws.append([])

        ws.append(["BOLSISTAS ATIVOS"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.row_dimensions[ws.max_row].height = 25

        if bolsistas:
            ws.append(["Nome", "CPF", "Nível", "Início Bolsa", "Fim Bolsa", "Status"])
            header_row = ws.max_row
            for col in range(1, 7):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            for nome, cpf, nivel, ini, fim, status in bolsistas:
                ini_br = self.converter_data_iso_para_br(ini)
                fim_br = self.converter_data_iso_para_br(fim)
                ws.append([nome, cpf, nivel, ini_br, fim_br, status])
                for col in range(1, 7):
                    ws.cell(row=ws.max_row, column=col).border = thin_border
        else:
            ws.append(["Nenhum bolsista ativo encontrado."])

        for col_idx in range(1, 7):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        wb.save(caminho)

    def gerar_pdf(self, caminho, modalidades, bolsistas, edital_filtro):
        doc = SimpleDocTemplate(caminho, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        if edital_filtro == "Todos":
            title = "Relatório de Gestão de Editais"
        else:
            title = f"Relatório do Edital {edital_filtro}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Modalidades", styles['Heading2']))
        if modalidades:
            data_mod = [["Nível", "Vagas", "Valor Mensal (R$)"]]
            for nivel, vagas, valor in modalidades:
                data_mod.append([nivel, str(vagas), f"R$ {valor:.2f}"])
            table_mod = Table(data_mod)
            table_mod.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            elements.append(table_mod)
        else:
            elements.append(Paragraph("Nenhuma modalidade encontrada.", styles['Normal']))
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Bolsistas Ativos", styles['Heading2']))
        if bolsistas:
            data_bols = [["Nome", "CPF", "Nível", "Início", "Fim", "Status"]]
            for nome, cpf, nivel, ini, fim, status in bolsistas:
                ini_br = self.converter_data_iso_para_br(ini)
                fim_br = self.converter_data_iso_para_br(fim)
                data_bols.append([nome, cpf, nivel, ini_br, fim_br, status])
            table_bols = Table(data_bols)
            table_bols.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('ALIGN', (3,1), (4,-1), 'CENTER')
            ]))
            elements.append(table_bols)
        else:
            elements.append(Paragraph("Nenhum bolsista ativo encontrado.", styles['Normal']))

        doc.build(elements)

if __name__ == "__main__":
    root = ThemedTk(theme="equilux")
    app = GestaoEditaisApp(root)
    root.mainloop()