"""
db.py - Módulo centralizado de acesso ao banco de dados.

Fornece funções para queries comuns e mantém a lógica SQL em um único lugar,
facilitando manutenção, testes e refatorações futuras.
"""

import sqlite3
import os
from datetime import datetime


DB_PATH = os.path.join(os.path.dirname(__file__), 'gestao_editais.db')


def get_connection():
    """Retorna uma conexão ao banco de dados."""
    return sqlite3.connect(DB_PATH)


# ===== EDITAIS =====

def obter_todos_editais():
    """Retorna lista de (id, numero_edital)."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, numero_edital FROM editais ORDER BY id")
    result = cursor.fetchall()
    conn.close()
    return result

    def ensure_db_schema():
        """Garante que as tabelas necessárias existam no banco.
        Executa CREATE TABLE IF NOT EXISTS para todas as tabelas usadas pela aplicação.
        """
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Tabela: editais
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS editais (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_edital TEXT NOT NULL,
                descricao TEXT,
                agencia_fomento TEXT NOT NULL,
                codigo_projeto TEXT,
                descricao_projeto TEXT
            )
        ''')

        # Tabela: modalidades
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS modalidades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                edital_id INTEGER NOT NULL,
                nivel TEXT NOT NULL,
                vagas INTEGER NOT NULL,
                valor_mensal REAL NOT NULL,
                FOREIGN KEY(edital_id) REFERENCES editais(id) ON DELETE CASCADE
            )
        ''')

        # Tabela: bolsistas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bolsistas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                edital_id INTEGER NOT NULL,
                processo_sei TEXT,
                nome TEXT NOT NULL,
                cpf TEXT,
                orientador TEXT,
                campus TEXT,
                programa TEXT,
                nivel TEXT NOT NULL,
                data_inicio_bolsa TEXT NOT NULL,
                meses_duracao INTEGER NOT NULL,
                data_fim_bolsa TEXT NOT NULL,
                previsao_defesa TEXT,
                email_bolsista TEXT,
                email_programa TEXT,
                status TEXT NOT NULL CHECK(status IN ('ativo', 'desligado', 'substituido')),
                FOREIGN KEY(edital_id) REFERENCES editais(id) ON DELETE CASCADE
            )
        ''')

        # Tabela: acompanhamento
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS acompanhamento (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bolsista_id INTEGER NOT NULL,
                referencia_mes TEXT NOT NULL,
                parcela INTEGER DEFAULT 0,
                requisicao_pagamento TEXT,
                observacoes TEXT,
                criado_em TEXT,
                UNIQUE(bolsista_id, referencia_mes),
                FOREIGN KEY(bolsista_id) REFERENCES bolsistas(id) ON DELETE CASCADE
            )
        ''')

        # Tabela: substituicoes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS substituicoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bolsista_original_id INTEGER NOT NULL,
                novo_bolsista_id INTEGER NOT NULL,
                data_substituicao TEXT NOT NULL,
                motivo TEXT,
                FOREIGN KEY(bolsista_original_id) REFERENCES bolsistas(id) ON DELETE CASCADE,
                FOREIGN KEY(novo_bolsista_id) REFERENCES bolsistas(id) ON DELETE CASCADE
            )
        ''')

        conn.commit()
        conn.close()


def obter_edital_por_id(edital_id):
    """Retorna dados de um edital."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto FROM editais WHERE id = ?", (edital_id,))
    result = cursor.fetchone()
    conn.close()
    return result


def criar_edital(numero, descricao, agencia, codigo_projeto=None, descricao_projeto=None):
    """Insere um novo edital."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO editais (numero_edital, descricao, agencia_fomento, codigo_projeto, descricao_projeto)
        VALUES (?, ?, ?, ?, ?)
    ''', (numero, descricao, agencia, codigo_projeto, descricao_projeto))
    edital_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return edital_id


# ===== BOLSISTAS =====

def obter_bolsistas_ativos():
    """Retorna lista de bolsistas com status 'ativo'."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, nome, nivel, data_inicio_bolsa, meses_duracao, data_fim_bolsa
        FROM bolsistas
        WHERE status = 'ativo'
        ORDER BY nome
    ''')
    result = cursor.fetchall()
    conn.close()
    return result


def obter_bolsista_por_id(bolsista_id):
    """Retorna dados completos de um bolsista."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT
            bolsistas.edital_id, numero_edital, processo_sei, nome, cpf, orientador,
            campus, programa, nivel, data_inicio_curso, data_inicio_bolsa,
            meses_duracao, previsao_defesa, email_bolsista, email_programa, status
        FROM bolsistas
        JOIN editais ON bolsistas.edital_id = editais.id
        WHERE bolsistas.id = ?
    ''', (bolsista_id,))
    result = cursor.fetchone()
    conn.close()
    return result


def criar_bolsista(edital_id, processo, nome, cpf, orientador, campus, programa,
                   nivel, data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
                   defesa_iso, email_bols, email_prog):
    """Insere um novo bolsista."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO bolsistas (
            edital_id, processo_sei, nome, cpf, orientador, campus, programa,
            nivel, data_inicio_curso, data_inicio_bolsa, meses_duracao, data_fim_bolsa,
            previsao_defesa, email_bolsista, email_programa, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        edital_id, processo, nome, cpf, orientador, campus, programa,
        nivel, data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
        defesa_iso, email_bols, email_prog, 'ativo'
    ))
    bolsista_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return bolsista_id


def atualizar_bolsista(bolsista_id, processo, nome, cpf, orientador, campus, programa,
                       data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
                       defesa_iso, email_bols, email_prog, status):
    """Atualiza dados de um bolsista."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE bolsistas SET
            processo_sei = ?, nome = ?, cpf = ?, orientador = ?, campus = ?, programa = ?,
            data_inicio_curso = ?, data_inicio_bolsa = ?, meses_duracao = ?, data_fim_bolsa = ?,
            previsao_defesa = ?, email_bolsista = ?, email_programa = ?, status = ?
        WHERE id = ?
    ''', (
        processo, nome, cpf, orientador, campus, programa,
        data_inicio_curso_iso, data_inicio_bolsa_iso, meses, data_fim_iso,
        defesa_iso, email_bols, email_prog, status, bolsista_id
    ))
    conn.commit()
    conn.close()


# ===== ACOMPANHAMENTO =====

def obter_acompanhamento_ativo(referencia=None):
    """Retorna acompanhamento para bolsistas ativos (filtro opcional por referência)."""
    conn = get_connection()
    cursor = conn.cursor()
    if referencia:
        cursor.execute('''
            SELECT a.id, b.id as bolsista_id, b.nome, b.cpf, b.nivel,
                   a.referencia_mes, a.parcela, a.requisicao_pagamento, a.observacoes
            FROM acompanhamento a
            JOIN bolsistas b ON a.bolsista_id = b.id
            WHERE b.status = 'ativo' AND a.referencia_mes = ?
            ORDER BY b.nome
        ''', (referencia,))
    else:
        cursor.execute('''
            SELECT a.id, b.id as bolsista_id, b.nome, b.cpf, b.nivel,
                   a.referencia_mes, a.parcela, a.requisicao_pagamento, a.observacoes
            FROM acompanhamento a
            JOIN bolsistas b ON a.bolsista_id = b.id
            WHERE b.status = 'ativo'
            ORDER BY a.referencia_mes DESC, b.nome
        ''')
    result = cursor.fetchall()
    conn.close()
    return result


def registrar_acompanhamento(bolsista_id, referencia_mes, parcela, requisicao, observacoes):
    """
    Insere ou atualiza acompanhamento.
    Usa ON CONFLICT com fallback para versões antigas do SQLite.
    """
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO acompanhamento (bolsista_id, referencia_mes, parcela, requisicao_pagamento, observacoes, criado_em)
            VALUES (?, ?, ?, ?, ?, datetime('now'))
            ON CONFLICT(bolsista_id, referencia_mes) DO UPDATE SET
                parcela=excluded.parcela,
                requisicao_pagamento=excluded.requisicao_pagamento,
                observacoes=excluded.observacoes,
                criado_em=datetime('now')
        ''', (bolsista_id, referencia_mes, parcela or 0, requisicao, observacoes))
    except sqlite3.OperationalError:
        # fallback
        cursor.execute('SELECT id FROM acompanhamento WHERE bolsista_id = ? AND referencia_mes = ?', (bolsista_id, referencia_mes))
        if cursor.fetchone():
            cursor.execute('''
                UPDATE acompanhamento SET parcela = ?, requisicao_pagamento = ?, observacoes = ?, criado_em = datetime('now')
                WHERE bolsista_id = ? AND referencia_mes = ?
            ''', (parcela or 0, requisicao, observacoes, bolsista_id, referencia_mes))
        else:
            cursor.execute('''
                INSERT INTO acompanhamento (bolsista_id, referencia_mes, parcela, requisicao_pagamento, observacoes, criado_em)
                VALUES (?, ?, ?, ?, ?, datetime('now'))
            ''', (bolsista_id, referencia_mes, parcela or 0, requisicao, observacoes))
    conn.commit()
    conn.close()


def obter_acompanhamento_para_csv():
    """Retorna registros de acompanhamento para exportação CSV."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT a.id, b.nome, b.cpf, b.nivel, a.referencia_mes, a.parcela, 
               a.requisicao_pagamento, a.observacoes, a.criado_em
        FROM acompanhamento a
        JOIN bolsistas b ON a.bolsista_id = b.id
        ORDER BY a.referencia_mes DESC, b.nome
    ''')
    result = cursor.fetchall()
    conn.close()
    return result


def prefill_mes_atual():
    """Pré-cria registros de acompanhamento para o mês atual para todos os bolsistas ativos."""
    from datetime import datetime
    ref = f"{datetime.now().year:04d}-{datetime.now().month:02d}"
    
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, data_inicio_bolsa FROM bolsistas WHERE status = 'ativo'")
    bolsistas = cursor.fetchall()
    
    inserted = 0
    for bolsista_id, inicio_iso in bolsistas:
        parcela = calcular_parcela(inicio_iso, ref)
        try:
            cursor.execute('''
                INSERT INTO acompanhamento (bolsista_id, referencia_mes, parcela, requisicao_pagamento, observacoes, criado_em)
                VALUES (?, ?, ?, NULL, NULL, datetime('now'))
                ON CONFLICT(bolsista_id, referencia_mes) DO NOTHING
            ''', (bolsista_id, ref, parcela))
            if cursor.rowcount == 1:
                inserted += 1
        except sqlite3.OperationalError:
            # fallback
            cursor.execute('SELECT id FROM acompanhamento WHERE bolsista_id = ? AND referencia_mes = ?', (bolsista_id, ref))
            if not cursor.fetchone():
                cursor.execute('INSERT INTO acompanhamento (bolsista_id, referencia_mes, parcela, requisicao_pagamento, observacoes, criado_em) VALUES (?, ?, ?, NULL, NULL, datetime(\'now\'))', (bolsista_id, ref, parcela))
                inserted += 1
    
    conn.commit()
    conn.close()
    return inserted, ref


def calcular_parcela(data_inicio_bolsa_iso, referencia_yyyy_mm):
    """Calcula número da parcela para uma referência (mês)."""
    try:
        inicio = datetime.strptime(data_inicio_bolsa_iso, '%Y-%m-%d')
        ref = datetime.strptime(referencia_yyyy_mm + '-01', '%Y-%m-%d')
        meses = (ref.year - inicio.year) * 12 + (ref.month - inicio.month)
        return meses + 1 if meses >= 0 else 0
    except Exception:
        return None


# ===== MODALIDADES =====

def obter_modalidades_por_edital(edital_id):
    """Retorna modalidades de um edital."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, nivel, vagas, valor_mensal FROM modalidades WHERE edital_id = ? ORDER BY nivel
    ''', (edital_id,))
    result = cursor.fetchall()
    conn.close()
    return result


def obter_edital_id_por_numero(numero_edital):
    """Retorna o ID de um edital pelo seu número."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM editais WHERE numero_edital = ?", (numero_edital,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None


def criar_modalidade(edital_id, nivel, vagas, valor_mensal):
    """Insere uma nova modalidade."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO modalidades (edital_id, nivel, vagas, valor_mensal)
        VALUES (?, ?, ?, ?)
    ''', (edital_id, nivel, vagas, valor_mensal))
    conn.commit()
    conn.close()


# ===== EXPORTAÇÃO EXCEL (COM FORMATAÇÃO) =====

def exportar_acompanhamento_para_excel(referencia_mes=None, caminho_saida=None):
    """
    Exporta acompanhamento para Excel com formatação profissional.
    
    Args:
        referencia_mes: Filtra por mês (ex: '2026-01'). Se None, exporta tudo.
        caminho_saida: Caminho do arquivo. Se None, salva como 'acompanhamento_YYYY-MM.xlsx'
    
    Returns:
        Caminho do arquivo gerado
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Obter dados
    conn = get_connection()
    cursor = conn.cursor()
    
    if referencia_mes:
        cursor.execute('''
            SELECT 
                e.numero_edital as 'Edital',
                b.processo_sei as 'SEI',
                b.cpf as 'CPF',
                b.nome as 'Nome',
                b.programa as 'Programa',
                b.campus as 'Campus',
                b.nivel as 'Nível',
                m.valor_mensal as 'Valor',
                b.data_inicio_bolsa as 'Início da Bolsa',
                a.referencia_mes as 'Referência',
                a.parcela as 'Parcela',
                a.requisicao_pagamento as 'Requisição',
                a.observacoes as 'Observações',
                a.criado_em as 'Data Criação'
            FROM acompanhamento a
            JOIN bolsistas b ON a.bolsista_id = b.id
            JOIN editais e ON b.edital_id = e.id
            LEFT JOIN modalidades m ON e.id = m.edital_id AND b.nivel = m.nivel
            WHERE a.referencia_mes = ?
            ORDER BY b.nome
        ''', (referencia_mes,))
    else:
        cursor.execute('''
            SELECT 
                e.numero_edital as 'Edital',
                b.processo_sei as 'SEI',
                b.cpf as 'CPF',
                b.nome as 'Nome',
                b.programa as 'Programa',
                b.campus as 'Campus',
                b.nivel as 'Nível',
                m.valor_mensal as 'Valor',
                b.data_inicio_bolsa as 'Início da Bolsa',
                a.referencia_mes as 'Referência',
                a.parcela as 'Parcela',
                a.requisicao_pagamento as 'Requisição',
                a.observacoes as 'Observações',
                a.criado_em as 'Data Criação'
            FROM acompanhamento a
            JOIN bolsistas b ON a.bolsista_id = b.id
            JOIN editais e ON b.edital_id = e.id
            LEFT JOIN modalidades m ON e.id = m.edital_id AND b.nivel = m.nivel
            ORDER BY a.referencia_mes DESC, b.nome
        ''')
    
    dados = cursor.fetchall()
    conn.close()
    
    if not dados:
        return None
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Acompanhamento"
    
    # Definir colunas (headers da query acima)
    headers = ['Edital', 'SEI', 'CPF', 'Nome', 'Programa', 'Campus', 'Nível', 
               'Valor', 'Início da Bolsa', 'Referência', 'Parcela', 'Requisição', 'Observações', 'Data Criação']
    
    # Escrever headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Escrever dados
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num, row_data in enumerate(dados, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Formatar valores monetários
            if col_num == 8:  # Coluna Valor
                if value:
                    cell.value = f"R$ {value:,.2f}".replace(',', '.')
                    cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Ajustar largura das colunas
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        if header in ['Nome', 'Programa', 'Observações']:
            ws.column_dimensions[col_letter].width = 25
        elif header in ['Requisição']:
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = 12
    
    # Definir altura das linhas
    ws.row_dimensions[1].height = 25
    for row_num in range(2, len(dados) + 2):
        ws.row_dimensions[row_num].height = 20
    
    # Salvar arquivo
    if caminho_saida is None:
        ref = referencia_mes if referencia_mes else f"{datetime.now().year:04d}-{datetime.now().month:02d}"
        caminho_saida = os.path.join(os.path.dirname(__file__), f"acompanhamento_{ref}.xlsx")
    
    wb.save(caminho_saida)
    return caminho_saida


def exportar_acompanhamento_mensal_automatico():
    """
    Exporta acompanhamento do mes atual para Excel automaticamente.
    Util para ser chamado pelo agendador (Task Scheduler).
    
    Returns:
        str: Caminho do arquivo gerado ou None se falhar
    """
    ref = f"{datetime.now().year:04d}-{datetime.now().month:02d}"
    caminho = exportar_acompanhamento_para_excel(referencia_mes=ref)
    
    if caminho:
        return caminho
    
    return None, ref, 0
